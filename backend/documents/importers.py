import calendar
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import Authority, DocumentDirection, DocumentTask, License, MiningObject

logger = logging.getLogger(__name__)

DOCUMENT_TITLE_COL = 8
DIRECTION_COL = 7
AVAILABILITY_COL = 9
RECEIVED_AT_COL = 10
COMMENT_COL = 11
LONG_COMMENT_COL = 120
CALENDAR_START_COL = 12
CALENDAR_END_COL = 119
MONTHS_IN_YEAR = 12
COLS_PER_MONTH = 9
COLS_PER_DECADE = 3

AUTHORITY_KEYWORDS = (
    ('роснедр', 'Роснедра'),
    ('росрыболов', 'Росрыболовство'),
    ('мпр', 'МПР'),
    ('ткз', 'ТКЗ'),
    ('ткр', 'ТКР'),
    ('куми', 'КУМИ'),
    ('санитар', 'Санитарно-эпидемиологическая служба'),
    ('экспертиз', 'Экспертиза'),
    ('суд', 'Суд'),
    ('админ', 'Администрация'),
)


@dataclass
class ImportStats:
    mining_objects_created: int = 0
    licenses_created: int = 0
    directions_created: int = 0
    authorities_created: int = 0
    tasks_created: int = 0
    tasks_updated: int = 0
    rows_skipped: int = 0
    warnings: list[str] = field(default_factory=list)


def import_document_plan(path, *, year=None, dry_run=False, update_existing=True):
    path = Path(path)
    source_year = year or _guess_year_from_path(path) or timezone.localdate().year
    stats = ImportStats()

    logger.info(
        'Document plan import started: path=%s year=%s dry_run=%s update_existing=%s',
        path,
        source_year,
        dry_run,
        update_existing,
    )

    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    _validate_worksheet(worksheet)

    with transaction.atomic():
        _import_worksheet(
            worksheet,
            stats=stats,
            source_year=source_year,
            update_existing=update_existing,
        )

        if dry_run:
            transaction.set_rollback(True)

    for warning in stats.warnings:
        logger.warning('Document plan import warning: %s', warning)

    logger.info(
        'Document plan import finished: path=%s year=%s dry_run=%s created=%s updated=%s skipped=%s',
        path,
        source_year,
        dry_run,
        stats.tasks_created,
        stats.tasks_updated,
        stats.rows_skipped,
    )
    return stats


def _validate_worksheet(worksheet):
    header_rows = []
    for row_number in range(1, worksheet.max_row + 1):
        marker = _text(worksheet.cell(row_number, 2).value)
        document_header = _text(worksheet.cell(row_number, DOCUMENT_TITLE_COL).value).lower()
        direction_header = _text(worksheet.cell(row_number, DIRECTION_COL).value).lower()
        if marker == '№':
            header_rows.append(row_number)
            if 'документ' not in document_header:
                raise ValueError(
                    f'Некорректная структура Excel: в строке {row_number}, колонке H ожидается "Документ".'
                )
            if 'направление' not in direction_header:
                raise ValueError(
                    f'Некорректная структура Excel: в строке {row_number}, колонке G ожидается "Направление".'
                )

    if not header_rows:
        raise ValueError('Некорректная структура Excel: не найдены строки заголовков с колонкой "№".')


def _import_worksheet(worksheet, *, stats, source_year, update_existing):
    section_name = None
    current_license_number = ''
    current_license_status = ''
    current_direction_name = ''

    for row_number in range(1, worksheet.max_row + 1):
        row_title = _text(worksheet.cell(row_number, 2).value)
        license_number = _text(worksheet.cell(row_number, 3).value)
        object_name = _text(worksheet.cell(row_number, 4).value)
        license_status = _text(worksheet.cell(row_number, 5).value)
        direction_name = _text(worksheet.cell(row_number, DIRECTION_COL).value)
        document_title = _text(worksheet.cell(row_number, DOCUMENT_TITLE_COL).value)

        if _is_section_row(row_title, license_number, document_title):
            section_name = row_title
            current_license_number = ''
            current_license_status = ''
            current_direction_name = ''
            logger.info('Document plan section detected: row=%s section=%s', row_number, section_name)
            continue

        if row_title == '№':
            stats.rows_skipped += 1
            continue

        if not document_title:
            stats.rows_skipped += 1
            continue

        if license_number:
            current_license_number = license_number
        if license_status:
            current_license_status = license_status
        if direction_name:
            current_direction_name = direction_name

        mining_object_name = object_name or section_name
        if not mining_object_name:
            _skip_row(stats, row_number, f'mining object not found for document "{document_title}"')
            continue

        mining_object, created = MiningObject.objects.get_or_create(
            name=_truncate(mining_object_name, 100),
        )
        if created:
            stats.mining_objects_created += 1

        license_object = None
        if current_license_number:
            license_object, created = License.objects.get_or_create(
                mining_object=mining_object,
                number=_truncate(current_license_number, 100),
                defaults={'status': _truncate(current_license_status, 150)},
            )
            if created:
                stats.licenses_created += 1
            elif current_license_status and license_object.status != current_license_status:
                license_object.status = _truncate(current_license_status, 150)
                license_object.save(update_fields=['status'])

        direction = None
        if current_direction_name:
            direction, created = DocumentDirection.objects.get_or_create(
                name=_truncate(current_direction_name, 100),
            )
            if created:
                stats.directions_created += 1

        comment = _build_comment(worksheet, row_number)
        authority = _get_authority(document_title, current_direction_name, comment, stats)
        received_at = _parse_received_at(worksheet.cell(row_number, RECEIVED_AT_COL).value)
        deadline = _parse_deadline(worksheet, row_number, source_year)
        is_available = _parse_availability(worksheet.cell(row_number, AVAILABILITY_COL).value)
        status = _guess_task_status(is_available, deadline)
        title = _truncate(document_title, 100)

        lookup = {
            'mining_object': mining_object,
            'license_object': license_object,
            'direction': direction,
            'title': title,
        }
        defaults = {
            'authority': authority,
            'status': status,
            'is_available': is_available,
            'received_at': received_at,
            'deadline': deadline,
            'comment': comment,
        }

        if update_existing:
            _, created = DocumentTask.objects.update_or_create(defaults=defaults, **lookup)
            if created:
                stats.tasks_created += 1
            else:
                stats.tasks_updated += 1
        else:
            exists = DocumentTask.objects.filter(**lookup).exists()
            if exists:
                _skip_row(stats, row_number, f'document already exists and --no-update is enabled: "{title}"')
                continue
            DocumentTask.objects.create(**lookup, **defaults)
            stats.tasks_created += 1


def _skip_row(stats, row_number, reason):
    stats.rows_skipped += 1
    message = f'row {row_number}: {reason}'
    stats.warnings.append(message)


def _is_section_row(row_title, license_number, document_title):
    return bool(row_title and row_title != '№' and not license_number and not document_title)


def _get_authority(document_title, direction_name, comment, stats):
    text = f'{document_title} {direction_name} {comment}'.lower()
    for keyword, authority_name in AUTHORITY_KEYWORDS:
        if keyword in text:
            authority, created = Authority.objects.get_or_create(name=authority_name)
            if created:
                stats.authorities_created += 1
            return authority
    return None


def _build_comment(worksheet, row_number):
    parts = []
    received_raw = worksheet.cell(row_number, RECEIVED_AT_COL).value
    short_comment = _text(worksheet.cell(row_number, COMMENT_COL).value)
    long_comment = _text(worksheet.cell(row_number, LONG_COMMENT_COL).value)

    if received_raw and not isinstance(received_raw, (datetime, date)):
        parts.append(f'Дата получения из Excel: {_text(received_raw)}')
    if short_comment:
        parts.append(short_comment)
    if long_comment:
        parts.append(long_comment)

    parts.append(f'Источник: {worksheet.title}, строка {row_number}')
    return '\n'.join(parts)


def _parse_received_at(value):
    if isinstance(value, datetime):
        return _aware(value)
    if isinstance(value, date):
        return _aware(datetime.combine(value, time.min))
    return None


def _parse_deadline(worksheet, row_number, source_year):
    text = ' '.join(
        _text(worksheet.cell(row_number, col).value)
        for col in (COMMENT_COL, LONG_COMMENT_COL)
    )

    deadline = _parse_deadline_from_text(text)
    if deadline:
        return deadline

    return _parse_deadline_from_calendar(worksheet, row_number, source_year)


def _parse_deadline_from_text(text):
    match = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})', text)
    if match:
        day, month, year = (int(part) for part in match.groups())
        if year < 100:
            year += 2000
        return _safe_datetime(year, month, day)

    match = re.search(r'до\s+(\d{4})\s*(?:г|год|года)?', text, flags=re.IGNORECASE)
    if match:
        return _safe_datetime(int(match.group(1)), 12, 31)

    return None


def _parse_deadline_from_calendar(worksheet, row_number, source_year):
    red_columns = [
        column
        for column in range(CALENDAR_START_COL, CALENDAR_END_COL + 1)
        if _is_red_cell(worksheet.cell(row_number, column))
    ]
    if not red_columns:
        return None

    return _datetime_from_calendar_column(red_columns[-1], source_year)


def _datetime_from_calendar_column(column, year):
    month_index = (column - CALENDAR_START_COL) // COLS_PER_MONTH
    decade_index = ((column - CALENDAR_START_COL) % COLS_PER_MONTH) // COLS_PER_DECADE

    if month_index < 0 or month_index >= MONTHS_IN_YEAR:
        return None

    month = month_index + 1
    if decade_index == 0:
        day = 10
    elif decade_index == 1:
        day = 20
    else:
        day = calendar.monthrange(year, month)[1]

    return _safe_datetime(year, month, day)


def _safe_datetime(year, month, day):
    try:
        return _aware(datetime(year, month, day, 23, 59))
    except ValueError:
        return None


def _is_red_cell(cell):
    color = cell.fill.fgColor
    return cell.fill.fill_type == 'solid' and color.type == 'rgb' and color.rgb == 'FFFF0000'


def _parse_availability(value):
    value = _text(value).lower()
    return value in {'есть', 'да', 'получено', 'имеется'}


def _guess_task_status(is_available, deadline):
    if is_available:
        return DocumentTask.Status.RECEIVED
    if deadline and deadline.date() < timezone.localdate():
        return DocumentTask.Status.OVERDUE
    return DocumentTask.Status.PLANNED


def _guess_year_from_path(path):
    match = re.search(r'(20\d{2})', path.name)
    if match:
        return int(match.group(1))
    return None


def _aware(value):
    if timezone.is_aware(value):
        return value
    return timezone.make_aware(value, timezone.get_current_timezone())


def _text(value):
    if value is None:
        return ''
    return str(value).replace('\n', ' ').strip()


def _truncate(value, max_length):
    value = _text(value)
    if len(value) <= max_length:
        return value
    return value[:max_length]
